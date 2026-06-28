#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
政治理论热词派🔥 —— 求是 Excel → data-qiushi.js
用法: python3 excel_to_qiushi_js.py 输入.xlsx [输出.js]
"""

import sys, json, re, os
import openpyxl

def excel_to_js(excel_path, js_path):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    headers = [str(c.value).strip() if c.value else '' for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    years = {}  # year -> {name, issues: {issue_id -> {title, cards[]}}}

    for row in ws.iter_rows(min_row=2, values_only=True):
        year    = str(row[col['年份']]).strip() if col.get('年份') is not None else ''
        iid     = str(row[col['期数ID']]).strip() if col.get('期数ID') is not None else ''
        ititle  = str(row[col.get('期数标题', 0)]).strip() if row[col.get('期数标题', 0)] else ''
        cid     = str(row[col.get('卡片ID', 0)]).strip() if row[col.get('卡片ID', 0)] else ''
        q       = str(row[col['问题']]).strip() if row[col.get('问题', 0)] else ''
        a       = str(row[col.get('答案', 0)]).strip() if row[col.get('答案', 0)] else ''
        ana     = str(row[col.get('解析', 0)]).strip() if row[col.get('解析', 0)] else ''
        exp     = str(row[col.get('拓展', 0)]).strip() if row[col.get('拓展', 0)] else ''

        if not year or not q:
            continue

        if year not in years:
            years[year] = {'name': f'{year}年求是', 'issues': {}}
        if iid not in years[year]['issues']:
            years[year]['issues'][iid] = {'title': ititle or f'{year}年求是', 'cards': []}

        card = {}
        card['id'] = cid or f"{iid}-{len(years[year]['issues'][iid]['cards'])+1:03d}"
        card['question'] = q
        card['answer'] = a
        if ana: card['analysis'] = ana
        if exp: card['expansion'] = exp
        years[year]['issues'][iid]['cards'].append(card)

    # 生成 JS
    js = "// 求是期刊数据 —— 由 excel_to_qiushi_js.py 自动生成\n// 生成时间：自动\n\n"
    js += "var cardData = cardData || {};\n"
    js += "cardData.qiushi = {\n"
    js += "    name: \"求是\",\n"
    js += "    years: {\n"

    for year in sorted(years.keys()):
        yd = years[year]
        js += f"        \"{year}\": {{\n"
        js += f"            name: {json.dumps(yd['name'], ensure_ascii=False)},\n"
        js += "            issues: [\n"

        # 按期数ID排序
        sorted_issues = sorted(yd['issues'].items(), key=lambda x: x[0])
        for iid, issue in sorted_issues:
            title = issue['title'] or f'{year}年求是'
            js += "                {\n"
            js += f"                    id: \"{iid}\",\n"
            js += f"                    title: {json.dumps(title, ensure_ascii=False)},\n"
            js += "                    cards: [\n"
            for c in issue['cards']:
                js += "                        " + json.dumps(c, ensure_ascii=False) + ",\n"
            js += "                    ]\n"
            js += "                },\n"

        js += "            ]\n"
        js += "        },\n"

    js += "    }\n"
    js += "};\n"

    with open(js_path, 'w', encoding='utf-8') as f:
        f.write(js)

    total = sum(len(i['cards']) for y in years.values() for i in y['issues'].values())
    print(f"✓ 已生成 {js_path}，共 {total} 张卡片，{len(years)} 年")

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("用法: python3 excel_to_qiushi_js.py 输入.xlsx [输出.js]")
        sys.exit(1)
    excel = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else 'data-qiushi.js'
    excel_to_js(excel, out)
