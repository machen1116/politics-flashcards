#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
更新求是数据（替换指定年份）- 正确版
用法: python3 update_qiushi_correct.py 新数据.xlsx --year 2026
"""

import sys, json, subprocess, os, argparse
import openpyxl

def update_qiushi_data(excel_path, target_year='2025'):
    # 1. 读取现有数据（从JSON）
    json_path = '/tmp/existing_qiushi.json'
    if not os.path.exists(json_path):
        print(f"✗ JSON文件不存在: {json_path}")
        print("请先运行 Node.js 脚本解析现有数据")
        sys.exit(1)
    
    with open(json_path, 'r', encoding='utf-8') as f:
        existing_data = json.load(f)
    
    print(f"现有数据年份: {sorted(existing_data['years'].keys())}")
    print(f"目标年份: {target_year}")
    
    # 2. 读取新Excel文件
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    headers = [str(c.value).strip() if c.value else '' for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}
    
    # 3. 解析新数据（只取目标年份）
    new_issues = {}  # issue_id -> {'title': ..., 'cards': []}
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        year = str(row[col['年份']]).strip() if row[col.get('年份', 0)] else ''
        iid = str(row[col.get('期数ID', col.get('文件ID', 0))]).strip() if row[col.get('期数ID', col.get('文件ID', 0))] else ''
        ititle = str(row[col.get('期数标题', col.get('文件标题', 0))]).strip() if row[col.get('期数标题', col.get('文件标题', 0))] else ''
        cid = str(row[col.get('卡片ID', 0)]).strip() if row[col.get('卡片ID', 0)] else ''
        q = str(row[col['问题']]).strip() if row[col.get('问题', 0)] else ''
        a = str(row[col.get('答案', 0)]).strip() if row[col.get('答案', 0)] else ''
        ana = str(row[col.get('解析', 0)]).strip() if row[col.get('解析', 0)] else ''
        exp = str(row[col.get('拓展', 0)]).strip() if row[col.get('拓展', 0)] else ''
        
        if not year or not q:
            continue
        
        # 只处理目标年份数据
        if year != target_year:
            continue
        
        if iid not in new_issues:
            new_issues[iid] = {'title': ititle or f'{target_year}年求是', 'cards': []}
        
        card = {
            'id': cid or f"{iid}-{len(new_issues[iid]['cards'])+1:03d}",
            'question': q,
            'answer': a
        }
        if ana:
            card['analysis'] = ana
        if exp:
            card['expansion'] = exp
        
        new_issues[iid]['cards'].append(card)
    
    # 4. 统计新数据
    new_total = sum(len(i['cards']) for i in new_issues.values())
    print(f"\n新{target_year}年数据: {len(new_issues)} 期, {new_total} 张卡片")
    
    # 5. 替换目标年份数据
    existing_data['years'][target_year] = {
        'name': f'{target_year}年求是',
        'issues': []
    }
    
    for iid in sorted(new_issues.keys()):
        issue = new_issues[iid]
        existing_data['years'][target_year]['issues'].append({
            'id': iid,
            'title': issue['title'],
            'cards': issue['cards']
        })
    
    # 6. 生成新的 JS 文件（使用正确格式）
    js_lines = []
    js_lines.append("// 求是期刊数据 —— 自动生成")
    js_lines.append("// 生成时间：2026-06-28\n")
    js_lines.append("var cardData = cardData || {};")
    js_lines.append("cardData.qiushi = {")
    js_lines.append('    name: "求是",')
    js_lines.append("    years: {")
    
    for year in sorted(existing_data['years'].keys()):
        yd = existing_data['years'][year]
        js_lines.append(f'        "{year}": {{')
        js_lines.append(f"            name: {json.dumps(yd['name'], ensure_ascii=False)},")
        js_lines.append("            issues: [")
        
        for issue in yd['issues']:
            js_lines.append("                {")
            js_lines.append(f'                    id: "{issue["id"]}",')
            js_lines.append(f"                    title: {json.dumps(issue['title'], ensure_ascii=False)},")
            js_lines.append("                    cards: [")
            
            for card in issue['cards']:
                # 使用 json.dumps 生成卡片的 JSON 格式
                card_json = json.dumps(card, ensure_ascii=False)
                js_lines.append(f"                        {card_json},")
            
            js_lines.append("                    ]")
            js_lines.append("                },")
        
        js_lines.append("            ]")
        js_lines.append("        },")
    
    js_lines.append("    }")
    js_lines.append("};")
    
    js_content = '\n'.join(js_lines)
    
    # 7. 写入文件
    output_path = '/Users/zhaochanghe/Desktop/政治理论热词/data-qiushi.js'
    
    # 备份
    if os.path.exists(output_path):
        backup_path = output_path.replace('.js', '-backup.js')
        import shutil
        shutil.copy2(output_path, backup_path)
        print(f"已备份: {backup_path}")
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(js_content)
    
    print(f"✓ 已更新 {output_path}")
    
    # 8. 验证
    total_all = 0
    for year in existing_data['years']:
        yd = existing_data['years'][year]
        total = sum(len(i['cards']) for i in yd['issues'])
        total_all += total
        print(f"  {year}: {len(yd['issues'])} 期, {total} 张卡片")
    
    print(f"\n总计: {total_all} 张卡片")

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='更新求是数据（替换指定年份）')
    parser.add_argument('excel', help='Excel 文件路径')
    parser.add_argument('--year', default='2025', help='要替换的年份（默认：2025）')
    
    args = parser.parse_args()
    
    # 先解析现有数据
    parse_script = '''
const fs = require('fs');
const vm = require('vm');
const code = fs.readFileSync('/Users/zhaochanghe/Desktop/政治理论热词/data-qiushi.js', 'utf-8');
const sandbox = { cardData: {} };
vm.createContext(sandbox);
vm.runInContext(code, sandbox);
const qs = sandbox.cardData.qiushi;
fs.writeFileSync('/tmp/existing_qiushi.json', JSON.stringify(qs, null, 2));
console.log('已解析现有数据');
'''
    
    with open('/tmp/parse_qiushi_correct.cjs', 'w', encoding='utf-8') as f:
        f.write(parse_script)
    
    result = subprocess.run(
        ['/Users/zhaochanghe/.workbuddy/binaries/node/versions/22.22.2/bin/node', '/tmp/parse_qiushi_correct.cjs'],
        capture_output=True,
        text=True
    )
    
    if result.returncode != 0:
        print(f"✗ 解析现有数据失败: {result.stderr}")
        sys.exit(1)
    
    update_qiushi_data(args.excel, args.year)
