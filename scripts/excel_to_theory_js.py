#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
政治理论热词派🔥 —— 理论基础知识 Excel → data-theory.js
用法: python3 excel_to_theory_js.py 输入.xlsx [输出.js]
"""

import sys, json, re, os
import openpyxl

VALID_TAGS = {'考点', '核心考点', '易错', '难点'}

def normalize_tag(v):
    if not v or not str(v).strip():
        return ''
    v = str(v).strip()
    if v in VALID_TAGS:
        return v
    # 模糊映射
    m = {'高频考点':'核心考点','重点':'核心考点','必考':'核心考点',
          '易错点':'易错','难点题':'难点','普通':'考点'}
    return m.get(v, '')

def excel_to_js(excel_path, js_path):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # 读表头
    headers = [str(c.value).strip() if c.value else '' for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    sections = {}  # section_id -> {name, cards[]}

    for row in ws.iter_rows(min_row=2, values_only=True):
        sid   = str(row[col['板块ID']]).strip() if col.get('板块ID') is not None else ''
        sname = str(row[col['板块名称']]).strip() if col.get('板块名称') is not None else ''
        cid   = str(row[col['卡片ID']]).strip() if col.get('卡片ID') is not None else ''
        q     = str(row[col['问题']]).strip() if row[col.get('问题', 0)] else ''
        a     = str(row[col['答案']]).strip() if row[col.get('答案', 0)] else ''
        ana   = str(row[col.get('解析', 0)]).strip() if row[col.get('解析', 0)] else ''
        exp   = str(row[col.get('拓展', 0)]).strip() if row[col.get('拓展', 0)] else ''
        tag   = normalize_tag(row[col.get('标签', 0)] if col.get('标签') is not None else '')
        diff  = row[col.get('难度(1-5)', 0)]
        if diff is not None:
            try: diff = int(diff)
            except: diff = 1
        else: diff = 1

        if not sid or not q:
            continue

        if sid not in sections:
            sections[sid] = {'id': sid, 'name': sname or sid, 'cards': []}

        card = {'id': cid or f"{sid}-{len(sections[sid]['cards'])+1:03d}"}
        card['question'] = q
        card['answer'] = a
        if ana: card['analysis'] = ana
        if exp: card['expansion'] = exp
        if tag: card['tagText'] = tag
        card['difficulty'] = diff
        sections[sid]['cards'].append(card)

    sections_list = [sections[sid] for sid in sections]

    # 生成 JS
    js = "// 理论基础知识数据 —— 由 excel_to_theory_js.py 自动生成\n// 生成时间：自动\n\n"
    js += "var cardData = cardData || {};\n"
    js += "cardData.theory = {\n"
    js += "    name: \"理论基础知识\",\n"
    js += "    sections: [\n"

    for sec in sections_list:
        js += "        {\n"
        js += f"            id: '{sec['id']}',\n"
        js += f"            name: {json.dumps(sec['name'], ensure_ascii=False)},\n"
        js += "            cards: [\n"
        for c in sec['cards']:
            js += "                " + json.dumps(c, ensure_ascii=False) + ",\n"
        js += "            ]\n"
        js += "        },\n"

    js += "    ]\n"
    js += "};\n"

    with open(js_path, 'w', encoding='utf-8') as f:
        f.write(js)
    print(f"✓ 已生成 {js_path}，共 {sum(len(s['cards']) for s in sections_list)} 张卡片，{len(sections_list)} 个板块")

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("用法: python3 excel_to_theory_js.py 输入.xlsx [输出.js]")
        sys.exit(1)
    excel = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else 'data-theory.js'
    excel_to_js(excel, out)
